import csv
import io
import logging
from datetime import datetime, date, timedelta
from decimal import Decimal
import os

from django.db import transaction
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from finance.models import Department
from email_notice.services import EmailNotificationService

from .filters import EvaluationRecordFilter
from .models import EvaluationRecord
from .serializers import EvaluationRecordSerializer, PersonnelSummarySerializer

logger = logging.getLogger(__name__)


class EvaluationRecordViewSet(viewsets.ModelViewSet):
    """考评记录视图集"""
    authentication_classes = [JWTAuthentication]
    queryset = EvaluationRecord.objects.select_related('department').all()
    serializer_class = EvaluationRecordSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = EvaluationRecordFilter
    search_fields = ['item_description', 'remarks', 'personnel', 'department__name', 'grade']
    ordering_fields = ['evaluation_date', 'created_at', 'total_score', 'bonus_score', 'deduction_score']
    ordering = ['-evaluation_date', '-created_at']

    def perform_create(self, serializer):
        """创建考评记录时发送邮箱通知"""
        instance = serializer.save()

        # 异步发送邮箱通知
        user_info = getattr(self.request.user, 'username', '系统') if hasattr(self.request, 'user') else '系统'
        EmailNotificationService.send_evaluation_operation_notification(
            'CREATE',
            evaluation_instance=instance,
            user_info=user_info
        )

    def perform_update(self, serializer):
        """更新考评记录时发送邮箱通知"""
        instance = serializer.save()

        # 异步发送邮箱通知
        user_info = getattr(self.request.user, 'username', '系统') if hasattr(self.request, 'user') else '系统'
        EmailNotificationService.send_evaluation_operation_notification(
            'UPDATE',
            evaluation_instance=instance,
            user_info=user_info
        )

    def perform_destroy(self, instance):
        """删除考评记录时发送邮箱通知"""
        # 异步发送邮箱通知
        user_info = getattr(self.request.user, 'username', '系统') if hasattr(self.request, 'user') else '系统'
        EmailNotificationService.send_evaluation_operation_notification(
            'DELETE',
            evaluation_instance=instance,
            user_info=user_info
        )

        # 执行删除
        instance.delete()

    @action(detail=False, methods=['get'], url_path='personnel-summary')
    def personnel_summary(self, request, *args, **kwargs):
        """获取人员汇总列表"""
        queryset = self.filter_queryset(self.get_queryset())
        
        # 按人员、部门、年级分组汇总
        summary_data = queryset.values('personnel', 'department__name', 'grade').annotate(
            total_bonus=Sum('bonus_score'),
            total_deduction=Sum('deduction_score'),
            bonus_count=Count('id', filter=Q(bonus_score__gt=0)),
            deduction_count=Count('id', filter=Q(deduction_score__gt=0)),
        ).order_by('department__name', 'personnel')
        
        # 计算总分
        result = []
        for item in summary_data:
            total_score = (item['total_bonus'] or Decimal('0')) - (item['total_deduction'] or Decimal('0'))
            result.append({
                'personnel': item['personnel'],
                'department_name': item['department__name'] or '',
                'grade': item['grade'] or '',
                'total_bonus': item['total_bonus'] or Decimal('0'),
                'total_deduction': item['total_deduction'] or Decimal('0'),
                'total_score': total_score,
                'bonus_count': item['bonus_count'],
                'deduction_count': item['deduction_count'],
            })
        
        serializer = PersonnelSummarySerializer(result, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='personnel-records')
    def personnel_records(self, request, *args, **kwargs):
        """获取某个人员的所有记录"""
        personnel_name = request.query_params.get('personnel')
        if not personnel_name:
            return Response({'detail': '缺少personnel参数'}, status=status.HTTP_400_BAD_REQUEST)
        queryset = self.get_queryset().filter(personnel=personnel_name).order_by('-evaluation_date')
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['delete'], url_path='delete-personnel')
    def delete_personnel(self, request, *args, **kwargs):
        """删除某个人员的所有记录"""
        personnel_name = request.query_params.get('personnel')
        department_name = request.query_params.get('department')
        grade = request.query_params.get('grade', '')
        
        if not personnel_name:
            return Response({'detail': '缺少personnel参数'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 构建查询条件
        filter_params = {'personnel': personnel_name}
        if department_name:
            filter_params['department__name'] = department_name
        if grade:
            filter_params['grade'] = grade
        
        # 查找要删除的记录
        queryset = self.get_queryset().filter(**filter_params)
        deleted_count = queryset.count()

        if deleted_count == 0:
            return Response({'detail': '未找到要删除的记录'}, status=status.HTTP_404_NOT_FOUND)

        # 收集删除的记录信息用于通知
        deleted_info = {
            'personnel': personnel_name,
            'department': department_name if department_name else '未指定',
            'grade': grade if grade else '未指定',
            'count': deleted_count
        }

        # 删除记录
        queryset.delete()

        # 异步发送邮箱通知
        user_info = getattr(request.user, 'username', '系统') if hasattr(request, 'user') else '系统'
        operation_description = f"删除人员考评记录 - {personnel_name}({deleted_info['department']}-{deleted_info['grade']})，共{deleted_count}条记录"

        EmailNotificationService.send_evaluation_operation_notification(
            'DELETE',
            evaluation_instance=None,
            user_info=user_info,
            operation_description=operation_description
        )

        return Response({
            'detail': f'成功删除 {deleted_count} 条记录',
            'deleted_count': deleted_count
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='export')
    def export_records(self, request, *args, **kwargs):
        """导出人员考评记录，每个人一个表格"""
        queryset = self.filter_queryset(self.get_queryset())
        if not queryset.exists():
            return Response({'detail': '暂无数据可导出'}, status=status.HTTP_400_BAD_REQUEST)

        # 创建Excel工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认工作表

        # 按人员分组
        personnel_groups = {}
        for record in queryset:
            personnel_key = f"{record.personnel}_{record.department.name if record.department else ''}"
            if personnel_key not in personnel_groups:
                personnel_groups[personnel_key] = {
                    'personnel': record.personnel,
                    'department_name': record.department.name if record.department else '',
                    'grade': record.grade or '',
                    'records': []
                }
            personnel_groups[personnel_key]['records'].append(record)

        # 样式定义
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        summary_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        summary_font = Font(bold=True)

        # 为每个人创建表格
        for idx, (key, group) in enumerate(personnel_groups.items(), 1):
            ws = wb.create_sheet(title=f"{group['personnel']}_{idx}")
            
            # 计算统计信息
            total_bonus = sum(record.bonus_score for record in group['records'])
            total_deduction = sum(record.deduction_score for record in group['records'])
            total_score = total_bonus - total_deduction
            bonus_count = sum(1 for record in group['records'] if record.bonus_score > 0)
            deduction_count = sum(1 for record in group['records'] if record.deduction_score > 0)

            # 基本信息行
            ws.append(['基本信息'])
            ws.append(['部门', group['department_name']])
            ws.append(['姓名', group['personnel']])
            ws.append(['年级', group['grade']])
            ws.append([])

            # 统计信息行
            ws.append(['统计信息'])
            ws.append(['总加分', f'{total_bonus:.2f}'])
            ws.append(['总扣分', f'{total_deduction:.2f}'])
            ws.append(['加分次数', bonus_count])
            ws.append(['扣分次数', deduction_count])
            ws.append([])

            # 记录表头
            headers = ['扣分/加分说明', '考评时间', '分值', '备注']
            ws.append(headers)

            # 设置表头样式
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 添加记录数据
            for record in sorted(group['records'], key=lambda x: x.evaluation_date, reverse=True):
                score_value = ''
                if record.bonus_score > 0:
                    score_value = f"+{record.bonus_score:.2f}"
                elif record.deduction_score > 0:
                    score_value = f"-{record.deduction_score:.2f}"
                
                ws.append([
                    record.item_description,
                    record.evaluation_date.strftime('%Y-%m-%d') if record.evaluation_date else '',
                    score_value,
                    record.remarks or '',
                ])

            # 设置列宽
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 30

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'人员考评记录_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        # 异步发送邮箱通知
        user_info = getattr(request.user, 'username', '系统') if hasattr(request, 'user') else '系统'
        operation_description = f"导出考评记录 - {len(personnel_groups)}名人员，共{queryset.count()}条记录"

        EmailNotificationService.send_evaluation_operation_notification(
            'CREATE',
            evaluation_instance=None,
            user_info=user_info,
            operation_description=operation_description
        )

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request, *args, **kwargs):
        """下载导入样表"""
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '人员导入样表'

        # 表头
        headers = ['部门', '年级', '姓名']
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 添加示例数据行
        ws.append(['程序部', '24', '张三'])
        ws.append(['Web部', '23', '李四'])


        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'人员导入样表_{timezone.now().strftime("%Y%m%d")}.xlsx'
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_records(self, request, *args, **kwargs):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        file_bytes = upload.read()
        records_data = []
        ext = os.path.splitext(upload.name or '')[1].lower()

        try:
            if ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
                records_data = self._read_excel(file_bytes)
            elif ext == '.csv' or not ext:
                decoded = file_bytes.decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                records_data = list(reader)
            else:
                decoded = file_bytes.decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                records_data = list(reader)
        except UnicodeDecodeError:
            return Response({'detail': '文件编码必须为UTF-8'}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if not records_data:
            return Response({'detail': '导入文件没有数据'}, status=status.HTTP_400_BAD_REQUEST)

        # 字段名映射：中文列名 -> 英文字段名
        field_mapping = {
            '部门': 'department_name',
            '所属部门': 'department_name',
            'department_name': 'department_name',
            '姓名': 'personnel_name',
            'personnel_name': 'personnel_name',
            '年级': 'grade',
            'grade': 'grade',
            '扣分/加分说明': 'item_description',
            'item_description': 'item_description',
            '加分': 'bonus_score',
            'bonus_score': 'bonus_score',
            '扣分': 'deduction_score',
            'deduction_score': 'deduction_score',
            '考评日期': 'evaluation_date',
            'evaluation_date': 'evaluation_date',
            '考评时间': 'evaluation_date',
            '备注': 'remarks',
            'remarks': 'remarks',
            '分值': 'score',
        }
        
        reader_fieldnames = records_data[0].keys() if isinstance(records_data[0], dict) else []
        # 获取实际存在的字段名（可能是中文或英文）
        available_field_keys = set()
        for field_name in reader_fieldnames:
            # 去除首尾空格后查找映射
            field_name_clean = field_name.strip() if field_name else ''
            mapped_key = field_mapping.get(field_name_clean)
            if mapped_key:
                available_field_keys.add(mapped_key)
        
        # 检查是样表格式（只有部门、年级、姓名）还是完整格式（有考评记录）
        # 样表格式：必须包含部门名和姓名，且不包含扣分/加分说明和考评日期
        is_template_format = (
            'department_name' in available_field_keys and
            'personnel_name' in available_field_keys and
            'item_description' not in available_field_keys and
            'evaluation_date' not in available_field_keys
        )
        
        if not is_template_format:
            # 完整格式需要检查必要的列
            required_fields = {'department_name', 'personnel_name', 'item_description', 
                              'evaluation_date'}
            missing_fields = required_fields - available_field_keys
            if missing_fields:
                # 将英文字段名转换为中文显示
                field_name_map = {
                    'department_name': '部门',
                    'personnel_name': '姓名',
                    'item_description': '扣分/加分说明',
                    'evaluation_date': '考评日期',
                }
                missing_display = [field_name_map.get(f, f) for f in sorted(missing_fields)]
                return Response(
                    {'detail': f'缺少必要的列: {", ".join(missing_display)}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        def get_field_value(row_dict, field_key):
            """从行数据中获取字段值，支持中英文列名"""
            # 收集所有可能的中文列名（去除首尾空格匹配）
            possible_keys = [field_key]  # 先尝试英文字段名
            
            # 找到所有映射到该字段的中文列名
            for chinese, english in field_mapping.items():
                if english == field_key:
                    possible_keys.append(chinese)
            
            # 尝试所有可能的键（包括去除空格后的键）
            for key in possible_keys:
                # 直接匹配
                if key in row_dict:
                    value = row_dict[key]
                    if value is not None:
                        return value
                # 去除空格后匹配（忽略键名和值的前后空格）
                for row_key in row_dict.keys():
                    if row_key and str(row_key).strip() == key.strip():
                        value = row_dict[row_key]
                        if value is not None:
                            return value
            
            return None

        records_to_create = []
        errors = []
        skipped_count = 0

        for idx, row in enumerate(records_data, start=2):
            try:
                if not isinstance(row, dict):
                    raise ValueError('数据格式不正确')

                # 根据 department_name 查找部门（支持中英文列名）
                department_name = get_field_value(row, 'department_name')
                if department_name is None:
                    # 显示实际存在的列名，帮助调试
                    available_keys = list(row.keys())
                    raise ValueError(f'部门名称不能为空。当前行的列名: {available_keys}')
                department_name = str(department_name).strip()
                if not department_name:
                    available_keys = list(row.keys())
                    raise ValueError(f'部门名称不能为空。当前行的列名: {available_keys}')
                department = Department.objects.filter(name=department_name).first()
                if not department:
                    raise ValueError(f'找不到部门: {department_name}')

                # personnel_name 直接作为字符串存储（支持中英文列名）
                personnel_name = str(get_field_value(row, 'personnel_name') or '').strip()
                if not personnel_name:
                    raise ValueError('人员名称不能为空')

                # 年级（可选）
                grade = str(get_field_value(row, 'grade') or '').strip()

                # 检查是否已存在相同姓名、年级、部门的人员记录
                existing_records = EvaluationRecord.objects.filter(
                    personnel=personnel_name,
                    grade=grade,
                    department=department
                ).exists()
                
                if existing_records:
                    skipped_count += 1
                    continue  # 跳过已存在的人员

                # 如果是样表格式（只有部门、年级、姓名），创建初始记录（总分为39）
                if is_template_format:
                    evaluation_date = timezone.now().date()
                    item_description = '初始分数'
                    bonus_score = Decimal('39')
                    deduction_score = Decimal('0')
                    remarks = ''
                else:
                    # 完整格式，从文件中读取
                    # 解析日期，支持多种格式（支持中英文列名）
                    evaluation_date_value = get_field_value(row, 'evaluation_date')
                    if evaluation_date_value is None:
                        raise ValueError('考评日期不能为空')
                    
                    # 如果已经是date类型，直接使用
                    if isinstance(evaluation_date_value, date):
                        evaluation_date = evaluation_date_value
                    else:
                        evaluation_date_str = str(evaluation_date_value).strip()
                        evaluation_date = None
                        for date_format in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
                            try:
                                evaluation_date = datetime.strptime(evaluation_date_str, date_format).date()
                                break
                            except ValueError:
                                continue
                        if evaluation_date is None:
                            raise ValueError(f'无法解析日期格式: {evaluation_date_str}')
                    
                    # 获取加分和扣分（支持中英文列名）
                    # 支持两种格式：1. 分别的加分/扣分列 2. 统一的分值列（+为加分，-为扣分）
                    score_value = get_field_value(row, 'score')
                    bonus_score = Decimal('0')
                    deduction_score = Decimal('0')
                    
                    if score_value:
                        # 从分值列解析
                        score_str = str(score_value).strip()
                        if score_str.startswith('+'):
                            bonus_score = Decimal(str(score_str[1:]) or '0')
                        elif score_str.startswith('-'):
                            deduction_score = Decimal(str(score_str[1:]) or '0')
                        else:
                            # 尝试解析为数字
                            try:
                                score_decimal = Decimal(score_str)
                                if score_decimal >= 0:
                                    bonus_score = score_decimal
                                else:
                                    deduction_score = abs(score_decimal)
                            except (ValueError, TypeError):
                                pass
                    else:
                        # 从分别的列读取
                        bonus_score = Decimal(str(get_field_value(row, 'bonus_score') or '0'))
                        deduction_score = Decimal(str(get_field_value(row, 'deduction_score') or '0'))
                    
                    item_description = str(get_field_value(row, 'item_description') or '').strip()
                    remarks = str(get_field_value(row, 'remarks') or '').strip()

                # total_score 会在 save() 方法中自动计算，不需要手动设置
                record = EvaluationRecord(
                    department=department,
                    personnel=personnel_name,
                    grade=grade,
                    item_description=item_description,
                    bonus_score=bonus_score,
                    deduction_score=deduction_score,
                    evaluation_date=evaluation_date,
                    remarks=remarks,
                )
                records_to_create.append(record)
            except Exception as exc:  # pylint: disable=broad-except
                errors.append(f'第 {idx} 行: {exc}')

        if errors:
            return Response({'detail': '导入失败', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        if not records_to_create:
            skip_msg = f'跳过 {skipped_count} 条已存在的记录' if skipped_count > 0 else ''
            return Response(
                {'detail': f'没有新数据需要导入。{skip_msg}'}, 
                status=status.HTTP_200_OK
            )

        with transaction.atomic():
            if not is_template_format:
                # 完整格式导入时，先删除所有记录（保持原有行为）
                EvaluationRecord.objects.all().delete()
            EvaluationRecord.objects.bulk_create(records_to_create)

        # 异步发送邮箱通知
        user_info = getattr(request.user, 'username', '系统') if hasattr(request, 'user') else '系统'
        operation_description = f"导入考评人员 - {len(records_to_create)}条记录"
        if is_template_format:
            operation_description += "（样表格式）"
        else:
            operation_description += "（完整格式）"

        EmailNotificationService.send_evaluation_operation_notification(
            'CREATE',
            evaluation_instance=None,
            user_info=user_info,
            operation_description=operation_description
        )

        skip_msg = f'，跳过 {skipped_count} 条已存在的记录' if skipped_count > 0 else ''
        return Response({
            'detail': f'成功导入 {len(records_to_create)} 条记录{skip_msg}'
        }, status=status.HTTP_200_OK)


    @staticmethod
    def _read_excel(file_bytes):
        workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [
            (str(cell).strip() if cell is not None else '').strip()
            for cell in rows[0]
        ]
        if not any(headers):
            raise ValueError('Excel表头为空')

        records = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if row is None:
                continue
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[col_idx] if col_idx < len(row) else None
                header_stripped = header.strip()
                header_key = header_stripped.lower()
                # 检查是否是日期列（支持中英文列名）
                is_date_column = (header_key == 'evaluation_date' or header_stripped == '考评日期')
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, date):
                    value = value.strftime('%Y-%m-%d')
                elif isinstance(value, (int, float)) and is_date_column:
                    value = EvaluationRecordViewSet._excel_date_to_iso(value, False)
                elif isinstance(value, Decimal):
                    value = str(value)
                elif isinstance(value, float):
                    formatted = format(value, 'f')
                    if '.' in formatted:
                        formatted = formatted.rstrip('0').rstrip('.')
                    value = formatted
                row_dict[header] = '' if value is None else value
            if any(value not in (None, '') for value in row_dict.values()):
                records.append(row_dict)
        return records

    @staticmethod
    def _excel_date_to_iso(excel_value, include_time):
        try:
            float_value = float(excel_value)
        except (TypeError, ValueError) as exc:
            raise ValueError('无法解析Excel日期') from exc
        base_date = datetime(1899, 12, 30)
        delta = timedelta(days=float_value)
        result = base_date + delta
        if include_time:
            return result.strftime('%Y-%m-%d %H:%M:%S')
        return result.strftime('%Y-%m-%d')
